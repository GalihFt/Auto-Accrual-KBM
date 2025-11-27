import streamlit as st
import pandas as pd
import numpy as np
import calendar
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import io

# ====================================================================
# I. KONFIGURASI APLIKASI STREAMLIT
# ====================================================================
st.set_page_config(
    page_title="Pemrosesan KBM Accrual",
    layout="wide"
)

# ====================================================================
# BAGIAN FUNGSI & ALGORITMA (TIDAK DIUBAH SAMA SEKALI)
# ====================================================================

# Fungsi untuk memuat file Excel dengan caching
@st.cache_data
def load_data(uploaded_file, sheet_name=None, header=None):
    """Membaca file Excel yang diunggah dengan caching."""
    try:
        if uploaded_file.name.endswith('.xls') or uploaded_file.name.endswith('.xlsx'):
            return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header)
        else:
            st.error("Tolong unggah file berformat .xls atau .xlsx.")
            return None
    except Exception as e:
        st.error(f"Error saat membaca file: {e}")
        return None

# Fungsi untuk memuat file Excel statis dengan caching
@st.cache_data
def load_static_file(file_path, sheet_name=None, header=None):
    """Membaca file Excel statis (list_COA.xlsx dan list_tarif.xlsx)"""
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)
    except FileNotFoundError:
        st.error(f"File penting tidak ditemukan: {file_path}. Pastikan file ini ada di direktori yang sama.")
        return None
    except Exception as e:
        st.error(f"Error saat memuat file statis {file_path}: {e}")
        return None

# Definisikan mapping di awal (konstanta)
list_bulan = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12
}
all_cabang_dict = {
    "BMS": "BANJARMASIN", "BPN": "BALIKPAPAN", "BTM": "BATAM", "MKS": "MAKASSAR", "MDN": "MEDAN",
    "PKB": "PEKANBARU", "SMD": "SAMARINDA", "SRG": "SORONG", "TRK": "TARAKAN", "MKW": "MANOKWARI",
    "MRK": "MERAUKE", "BTJ": "BATULICIN", "BUW": "BAU - BAU", "BER": "BERAU", "BIK": "BIAK",
    "BIT": "BITUNG", "FAK": "FAK - FAK", "KTG": "KETAPANG", "KNG": "KAIMANA", "NBI": "NABIRE",
    "NNK": "NUNUKAN", "SPT": "SAMPIT", "SRI": "SERUI", "TRT": "TERNATE", "TMK": "TIMIKA",
    "TUA": "TUAL", "PNK": "PONTIANAK", "PBG": "PALEMBANG", "AMB": "AMBON", "GOR": "GORONTALO",
    "PALU": "PALU", "PDG": "PADANG", "KDI": "KENDARI", "SMG": "SEMARANG", "BKU": "BUNGKU",
    "SBY": "SURABAYA"
}
list_port_full = ["BMS","MKS","MDN","PKB","SMD","SRG","TRK","MRK","BTJ","BER","BIT","SPT","AMB","GOR","PALU", "SBY"]
list_cy_full = ["BPN","BTM","MKW","BUW","BIK","FAK","KTG","KNG","NBI","NNK","SRI","TRT","TMK","TUA","PNK","PBG","PDG","KDI","SMG","BKU"]


# Pindahkan logika utama ke dalam fungsi agar dapat dipanggil oleh tombol
def run_processing(uploaded_file, input_bulan, input_tahun, selected_cabang, list_bulan, all_cabang_dict, list_port_full, list_cy_full):
    if not uploaded_file:
        st.warning("Tolong unggah file DATA_KBM.xls terlebih dahulu.")
        return None, None
    if not selected_cabang:
        st.warning("Tolong pilih setidaknya satu cabang.")
        return None, None
    if not input_tahun.isdigit() or len(input_tahun) != 4:
        st.error("Format Tahun tidak valid.")
        return None, None
    
    try:
        # Pengecekan dan konversi input
        bulan_num = list_bulan[input_bulan.lower()]
        tahun_num = int(input_tahun)

        # A. PERHITUNGAN TANGGAL & REGEX (sama seperti notebook)
        last_day = calendar.monthrange(tahun_num, bulan_num)[1]
        tanggal_akhir = f"{last_day:02d}/{bulan_num:02d}/{tahun_num}"
        
        if bulan_num == 12:
            bulan_next = 1
            tahun_next = tahun_num + 1
        else:
            bulan_next = bulan_num + 1
            tahun_next = tahun_num
            
        kode_yymm = f"{str(tahun_next)[2:]}{bulan_next:02d}"
        regex_kode = f"/{kode_yymm}/"

        st.info(f"Memproses data hingga **{tanggal_akhir}**. Mencari kode bulan berikutnya (JMH) **{kode_yymm}**.")
        
        # Penentuan cabang
        list_port = [x for x in list_port_full if x in selected_cabang]
        list_cy   = [x for x in list_cy_full if x in selected_cabang]
        full_list = selected_cabang

        # B. LOAD DATA STATIS (list_COA dan price_list)
        list_COA = pd.read_excel("list_COA.xlsx")
        price_list = pd.read_excel("list_tarif.xlsx")
        
        if list_COA is None or price_list is None:
            return None, None

        list_COA.index = list_COA["Nama Kegiatan"]
        list_COA.drop(columns=["Nama Kegiatan"], inplace=True)
        
        price_list.rename(columns={"CABANG": "key_cabang"}, inplace=True) 
        price_list = price_list[["key_cabang", "STVDR", "HAULAGE", "LOLO BM"]]

        # C. LOAD DATA KBM (semua sheet)
        all_sheets = load_data(uploaded_file, sheet_name=None, header=None)
        if all_sheets is None:
            return None, None
        
        sheets_list = list(all_sheets.values())
        
        # --- Fungsi Pembersihan DataFrame (diekstrak untuk DRY) ---
        def clean_df(df, is_fb=False):
            df_clean = df.dropna(thresh=5)
            df_clean.columns = df_clean.iloc[0]
            df_clean = df_clean[1:]
            df_clean = df_clean.reset_index(drop=True)
            
            if is_fb: # Format Baru (FB)
                df_clean["Qty Angkatan"] = pd.to_numeric(df_clean["Qty Angkatan"], errors='coerce')
                df_clean["vesvoy"] = df_clean["Vessel Id"] + " " + df_clean["Voyage No"]
            else: # Format Lama (FL)
                df_clean["Sub Total"] = pd.to_numeric(df_clean["Sub Total"], errors='coerce')
                df_clean["vesvoy"] = df_clean["Vessel Id To"] + " " + df_clean["Voyage No To"]
                
            df_clean["Status_KBM"] = None
            return df_clean

        # --- Pembersihan dan Penggabungan Data FL (Format Lama: Sheet 2, 0, 1) ---
        FL_clean = clean_df(sheets_list[2])
        FL1_clean = clean_df(sheets_list[0])
        FL2_clean = clean_df(sheets_list[1])
        
        # Filtering NEXT_MONTH_DOC untuk FL1 dan FL2
        FL1_clean.loc[
            FL1_clean["No Dokumen"].astype(str).str.contains(regex_kode, regex=True, na=False),
            "Status_KBM"
        ] = "NEXT_MONTH_DOC"
        
        FL2_clean.loc[
            FL2_clean["No Dokumen"].astype(str).str.contains(regex_kode, regex=True, na=False),
            "Status_KBM"
        ] = "NEXT_MONTH_DOC"

        # Terapkan status NO_DOC ke FL_clean (Sheet 2)
        FL_clean.loc[
            (FL_clean["Jenis Dokumen"] == "-") &
            (FL_clean["Port Id"].isin(list_port)) &
            (FL_clean["Status"].isin(["EMPTY", "-"])),
            "Status_KBM"
        ] = "NO_DOC_PORT"

        FL_clean.loc[
            (FL_clean["Jenis Dokumen"] == "-") &
            (FL_clean["Port Id"].isin(list_cy)) &
            (FL_clean["Status"].isin(["EMPTY", "-", "FULL"])),
            "Status_KBM"
        ] = "NO_DOC_CY"
        
        # Terapkan status NEXT_MONTH_DOC ke FL_clean (Sheet 2)
        FL_clean.loc[
            FL_clean["No Dokumen"].astype(str).str.contains(regex_kode, regex=True, na=False),
            "Status_KBM"
        ] = "NEXT_MONTH_DOC"

        # Gabungkan FL yang memiliki status
        FL_clean = pd.concat([
            FL_clean,
            FL1_clean[FL1_clean["Status_KBM"].notna()],
            FL2_clean[FL2_clean["Status_KBM"].notna()]
        ], ignore_index=True)


        # --- Pembersihan Data FB (Format Baru: Sheet 3) ---
        FB_clean = clean_df(sheets_list[3], is_fb=True)

        # Terapkan Status KBM ke FB_clean
        FB_clean.loc[
            ((FB_clean["Id Document"] == "-") | (FB_clean["Id Document"].isna()))
            & FB_clean["Port Id"].isin(list_port)
            & FB_clean["Type Size Name"].astype(str).str.contains("MT", na=False),
            "Status_KBM"
        ] = "NO_DOC_PORT"

        FB_clean.loc[
            ((FB_clean["Id Document"] == "-") | (FB_clean["Id Document"].isna()))
            & FB_clean["Port Id"].isin(list_cy)
            & FB_clean["Type Size Name"].astype(str).str.contains("MT|FL|-", na=False),
            "Status_KBM"
        ] = "NO_DOC_CY"

        FB_clean.loc[
            FB_clean["Id Document"].astype(str).str.contains(regex_kode, regex=True, na=False),
            "Status_KBM"
        ] = "NEXT_MONTH_DOC"

        # Merge dengan Price List
        FB_clean["key_cabang"] = FB_clean["Port Id"] + " " + FB_clean["Type Size Name"]
        FB_clean = pd.merge(
            left=FB_clean, 
            right=price_list, 
            on='key_cabang', 
            how='left' 
        )
        
        # Hitung biaya
        FB_clean["STVDR"] = FB_clean["Qty Angkatan"] * FB_clean["STVDR"]
        FB_clean["HAULAGE"] = FB_clean["Qty Angkatan"] * FB_clean["HAULAGE"]
        FB_clean["LOLO BM"] = FB_clean["Qty Angkatan"] * FB_clean["LOLO BM"]
        FB_clean.drop(columns=["key_cabang"], inplace=True)
        
        # Cek status dokumen
        dokumen_set = set(FL_clean["No Dokumen"].astype(str).unique())
        
        def check_status(id_doc_string):
            if pd.isna(id_doc_string) or str(id_doc_string).strip() == "-":
                return None
            dok_list = [d.strip() for d in str(id_doc_string).split(",")]
            return "hide" if any(d in dokumen_set for d in dok_list) else "show"

        FB_clean["Status_dokumen"] = FB_clean["Id Document"].apply(check_status)

        # =========================================================================
        # D. PEMBENTUKAN OUTPUT DETAIL (dfs_FL dan dfs_FB)
        # =========================================================================
    
        dfs_FL = {}
        dfs_FB = {}

        # 1. Definisi Kolom Tampilan (Eksplisit)
        # --------------------------------------
        kolom_tampilan_fl = [
            'Id KBM', 'Tgl KBM', 'Port Id', 'TD Month', 'Vessel Id To', 'Voyage No To', 
            'Nama Kegiatan', 'Jenis', 'Ukuran', 'Status', 'Jumlah Container', 'Biaya', 
            'Sub Total', 'Jenis Dokumen', 'No Dokumen', 'Tgl Create Documen', 
            'Tgl Kasir Documen', 'Created By', 'Port Id From', 'Port Id To', 'Kode ACC', 
            'Supplier', 'Id BS Penyelesaian', 'Tgl Kasir Id BS Penyelesaian', 'Id BKM', 
            'vesvoy', 'Status_KBM'
        ]

        kolom_tampilan_fb = [
            'No.', 'Vessel Id', 'Voyage No', 'TD', 'Port Id', 'Load Port', 'Disc Port', 
            'Id Session', 'Vessel Id From', 'Voyage No From', 'Vessel Id To', 'Voyage No To', 
            'Port Id From', 'Port Id To', 'Type Size Name', 'Qty Angkatan', 'Nama Vendor', 
            'ETS Status', 'Activity System Name', 'Id KBM', 'Tanggal', 'Id Document', 
            'vesvoy', 'Status_KBM', 'STVDR', 'HAULAGE', 'LOLO BM', 'Status_dokumen'
        ]

        # 2. Loop per Port
        # --------------------------------------
        for port in full_list:
            
            # --- PROSES FORMAT LAMA (FL) ---
            # A. Filter Data per Port
            df_port_fl = FL_clean[FL_clean["Port Id"] == port].copy()
            
            # B. SELEKSI KOLOM (PENTING: Agar kolom tidak melebar)
            df_port_fl = df_port_fl.reindex(columns=kolom_tampilan_fl)
            
            # C. Split Data (No Doc vs Next Month)
            df_no_doc_fl = df_port_fl[df_port_fl["Status_KBM"].astype(str).str.contains("NO_DOC", na=False)].copy()
            df_next_fl = df_port_fl[df_port_fl["Status_KBM"] == "NEXT_MONTH_DOC"].copy()
            
            # D. Buat Header & Spacer (Mengikuti kolom_tampilan_fl)
            header1_fl = pd.DataFrame({"Id KBM": [f"FORMAT LAMA - NO DOCUMENT"]})
            header1_fl = header1_fl.reindex(columns=kolom_tampilan_fl, fill_value=None)
            
            header2_fl = pd.DataFrame({"Id KBM": [f"FORMAT LAMA - JMH NEXT MONTH"]})
            header2_fl = header2_fl.reindex(columns=kolom_tampilan_fl, fill_value=None)
            
            spacer_fl = pd.DataFrame(
                [[None] * len(kolom_tampilan_fl)], 
                columns=kolom_tampilan_fl
            )
            
            # E. Gabungkan menjadi satu DataFrame
            dfs_FL[port] = pd.concat(
                [header1_fl, df_no_doc_fl, spacer_fl, header2_fl, df_next_fl], 
                ignore_index=True
            )

            # --- PROSES FORMAT BARU (FB) ---
            # A. Filter Data per Port
            df_port_fb = FB_clean[FB_clean["Port Id"] == port].copy()
            
            # B. SELEKSI KOLOM (PENTING: Agar kolom tidak melebar dan membuang sisa merge)
            df_port_fb = df_port_fb.reindex(columns=kolom_tampilan_fb)

            # C. Split Data (No Doc vs Next Month)
            df_no_doc_fb = df_port_fb[df_port_fb["Status_KBM"].astype(str).str.contains("NO_DOC", na=False)].copy()
            # Set Status_dokumen jadi NaN untuk bagian NO_DOC (sesuai logika notebook)
            df_no_doc_fb["Status_dokumen"] = np.nan 
            
            df_next_fb = df_port_fb[df_port_fb["Status_KBM"] == "NEXT_MONTH_DOC"].copy()
            # Kosongkan nilai biaya untuk bagian NEXT MONTH
            df_next_fb["STVDR"] = np.nan
            df_next_fb["HAULAGE"] = np.nan
            df_next_fb["LOLO BM"] = np.nan
            
            # D. Buat Header & Spacer (Mengikuti kolom_tampilan_fb)
            header1_fb = pd.DataFrame({"Vessel Id": ["FORMAT BARU - NO DOCUMENT"]})
            header1_fb = header1_fb.reindex(columns=kolom_tampilan_fb, fill_value=None)
            
            header2_fb = pd.DataFrame({"Vessel Id": ["FORMAT BARU - JMH NEXT MONTH"]})
            header2_fb = header2_fb.reindex(columns=kolom_tampilan_fb, fill_value=None)
            
            spacer_fb = pd.DataFrame(
                [[None] * len(kolom_tampilan_fb)], 
                columns=kolom_tampilan_fb
            )

            # E. Gabungkan menjadi satu DataFrame
            dfs_FB[port] = pd.concat(
                [header1_fb, df_no_doc_fb, spacer_fb, header2_fb, df_next_fb], 
                ignore_index=True
            )

        # E. PEMBENTUKAN JURNAL NO JMH (FL)
        FL_NO_JMH = FL_clean[FL_clean["Status_KBM"].astype(str).str.contains("NO_DOC", na=False)].copy()
        FL_NO_JMH["Keperluan"] = FL_NO_JMH["vesvoy"] + " " + FL_NO_JMH["Nama Kegiatan"] + " " + FL_NO_JMH["Ukuran"] + " " + FL_NO_JMH["Status"].astype(str)
        FL_NO_JMH = FL_NO_JMH[["Port Id", "Keperluan", "vesvoy", "Sub Total", "Kode ACC"]]

        FL_NO_JMH = FL_NO_JMH.join(list_COA, on="Kode ACC", how="left")
        FL_NO_JMH_FINAL = FL_NO_JMH[["Port Id", "Keperluan", "vesvoy", "Sub Total", "COA"]].copy()
        FL_NO_JMH_FINAL["Kredit"] = 0
        FL_NO_JMH_FINAL["COA-K"] = "-"
        FL_NO_JMH_FINAL["KODE"] = 1
        FL_NO_JMH_FINAL.rename(columns={"Sub Total": "Debit"}, inplace=True)
        FL_NO_JMH_FINAL.reset_index(drop=True, inplace= True)

        # Proses Accrual FL
        df_JMH_FL = FL_NO_JMH_FINAL.copy()
        total_per_port = df_JMH_FL.groupby("Port Id")["Debit"].sum().reset_index(name="Kredit")
        accrue_rows = total_per_port.copy()
        accrue_rows["Keperluan"] = f"ACCRUE {input_bulan} {input_tahun}"
        accrue_rows["vesvoy"] = "-"
        accrue_rows["Debit"] = 0
        accrue_rows["COA"] = "-"
        accrue_rows["COA-K"] = "3XX.01.12"
        accrue_rows["KODE"] = 3
        accrue_rows = accrue_rows[df_JMH_FL.columns]
        FL_NO_JMH_FINAL_FIX = pd.concat([df_JMH_FL, accrue_rows], ignore_index=True)
        FL_NO_JMH_FINAL_FIX = FL_NO_JMH_FINAL_FIX.sort_values(["Port Id", "KODE"]).reset_index(drop=True)

        FL_NO_JMH_FINAL_FIX["tanggal"], FL_NO_JMH_FINAL_FIX["nama_jurnal"], FL_NO_JMH_FINAL_FIX["A"], FL_NO_JMH_FINAL_FIX["kolom_1"] = "", "", "", ""
        first_idx = FL_NO_JMH_FINAL_FIX.groupby("Port Id").head(1).index

        for idx in first_idx:
            FL_NO_JMH_FINAL_FIX.loc[idx, "tanggal"] = tanggal_akhir
            FL_NO_JMH_FINAL_FIX.loc[idx, "nama_jurnal"] = f"ACCRUE {input_bulan} {input_tahun}"
            FL_NO_JMH_FINAL_FIX.loc[idx, "A"] = "A"
            FL_NO_JMH_FINAL_FIX.loc[idx, "kolom_1"] = 1

        FL_NO_JMH_FINAL_FIX = FL_NO_JMH_FINAL_FIX[(FL_NO_JMH_FINAL_FIX["Debit"] + FL_NO_JMH_FINAL_FIX["Kredit"]) != 0]
        cols_to_blank = ["Port Id", "tanggal", "nama_jurnal", "A", "kolom_1"]
        for col in cols_to_blank:
            FL_NO_JMH_FINAL_FIX.loc[~FL_NO_JMH_FINAL_FIX.index.isin(first_idx), col] = ""

        # F. PEMBENTUKAN JURNAL NO JMH (FB)
        FB_no_JMH = FB_clean[FB_clean["Status_KBM"].astype(str).str.contains("NO_DOC", na=False)].copy()
        FB_no_JMH["Keperluan"] = FB_no_JMH["vesvoy"] + " " + FB_no_JMH["Type Size Name"]
        FB_no_JMH = FB_no_JMH[["Port Id", "Keperluan", "vesvoy", "STVDR", "HAULAGE", "LOLO BM"]]

        FB_NO_JMH_STVDR = FB_no_JMH[["Port Id", "Keperluan", "vesvoy", "STVDR"]].copy()
        FB_NO_JMH_STVDR["COA"] = np.where((FB_NO_JMH_STVDR["Keperluan"].str.contains("MT", na=False)), "7XX.03.01.02.04", "7XX.03.01.02.03").copy()
        FB_NO_JMH_STVDR.rename(columns={"STVDR": "Debit"}, inplace=True)

        FB_NO_JMH_HAULAGE = FB_no_JMH[["Port Id", "Keperluan", "vesvoy", "HAULAGE"]].copy()
        FB_NO_JMH_HAULAGE["COA"] = "7XX.18.01"
        FB_NO_JMH_HAULAGE.rename(columns={"HAULAGE": "Debit"}, inplace=True)

        FB_NO_JMH_LOLO = FB_no_JMH[["Port Id", "Keperluan", "vesvoy", "LOLO BM"]].copy()
        FB_NO_JMH_LOLO["COA"] = "7XX.04.02"
        FB_NO_JMH_LOLO.rename(columns={"LOLO BM": "Debit"}, inplace=True)

        FB_NO_JMH_FINAL = pd.concat([FB_NO_JMH_STVDR, FB_NO_JMH_HAULAGE, FB_NO_JMH_LOLO], ignore_index=True)
        
        # Proses Accrual FB
        FB_NO_JMH_FINAL["Kredit"] = 0
        FB_NO_JMH_FINAL["COA-K"] = "-"
        FB_NO_JMH_FINAL["KODE"] = 1
        FB_NO_JMH_FINAL.reset_index(drop=True, inplace=True)
        df_JMH_FB = FB_NO_JMH_FINAL.copy()

        total_per_port_FB = df_JMH_FB.groupby("Port Id")["Debit"].sum().reset_index(name="Kredit")
        accrue_FB = total_per_port_FB.copy()
        accrue_FB["Keperluan"] = f"ACCRUE XYZ {input_bulan} {input_tahun}"
        accrue_FB["vesvoy"] = "-"
        accrue_FB["Debit"] = 0
        accrue_FB["COA"] = "-"
        accrue_FB["COA-K"] = "3XX.01.12"
        accrue_FB["KODE"] = 3
        accrue_FB = accrue_FB[df_JMH_FB.columns]
        
        FB_NO_JMH_FINAL_FIX = pd.concat([df_JMH_FB, accrue_FB], ignore_index=True)
        FB_NO_JMH_FINAL_FIX = FB_NO_JMH_FINAL_FIX.sort_values(["Port Id", "KODE"]).reset_index(drop=True)
        
        FB_NO_JMH_FINAL_FIX["tanggal"], FB_NO_JMH_FINAL_FIX["nama_jurnal"], FB_NO_JMH_FINAL_FIX["A"], FB_NO_JMH_FINAL_FIX["kolom_1"] = "", "", "", ""
        first_idx_FB = FB_NO_JMH_FINAL_FIX.groupby("Port Id").head(1).index

        for idx in first_idx_FB:
            FB_NO_JMH_FINAL_FIX.loc[idx, "tanggal"] = tanggal_akhir
            FB_NO_JMH_FINAL_FIX.loc[idx, "nama_jurnal"] = f"ACCRUE XYZ {input_bulan} {input_tahun}"
            FB_NO_JMH_FINAL_FIX.loc[idx, "A"] = "A"
            FB_NO_JMH_FINAL_FIX.loc[idx, "kolom_1"] = 1
        
        FB_NO_JMH_FINAL_FIX = FB_NO_JMH_FINAL_FIX[(FB_NO_JMH_FINAL_FIX["Debit"] + FB_NO_JMH_FINAL_FIX["Kredit"]) != 0]

        cols_to_blank = ["Port Id", "tanggal", "nama_jurnal", "A", "kolom_1"]
        for col in cols_to_blank:
            FB_NO_JMH_FINAL_FIX.loc[~FB_NO_JMH_FINAL_FIX.index.isin(first_idx_FB), col] = ""

        # Urutan Kolom Jurnal
        ordered_cols = ["tanggal", "Port Id", "nama_jurnal", "A", "kolom_1", "Keperluan", "KODE", "vesvoy", "Debit", "Kredit", "COA", "COA-K"]
        FL_NO_JMH_FINAL_FIX = FL_NO_JMH_FINAL_FIX[ordered_cols]
        FB_NO_JMH_FINAL_FIX = FB_NO_JMH_FINAL_FIX[ordered_cols]

        # G. FINAL JURNAL DAN LIST JMH
        FULL_NO_JMH = pd.concat([FL_NO_JMH_FINAL_FIX, FB_NO_JMH_FINAL_FIX], ignore_index=True)
        mask = (FULL_NO_JMH["Port Id"] == "SBY") & (FULL_NO_JMH["A"] == "A")
        FULL_NO_JMH.loc[mask, "A"] = "B"
        FULL_NO_JMH["Port Id"] = FULL_NO_JMH["Port Id"].map(all_cabang_dict)

        # Pembuatan List JMH Gabungan
        list_JMH_FL = FL_clean[FL_clean["Status_KBM"] == "NEXT_MONTH_DOC"][["Port Id","No Dokumen", "vesvoy"]].drop_duplicates()
        list_JMH_FL["sumber"] = "FL"

        list_JMH_FB = FB_clean.assign(
            Id_Document_Split = FB_clean["Id Document"].astype(str).str.split(",")
        ).explode("Id_Document_Split")

        list_JMH_FB["Id_Document_Split"] = list_JMH_FB["Id_Document_Split"].str.strip()
        list_JMH_FB = list_JMH_FB[["Port Id","Id_Document_Split", "vesvoy"]].drop_duplicates()
        list_JMH_FB = list_JMH_FB[list_JMH_FB["Id_Document_Split"].notna()]
        list_JMH_FB = list_JMH_FB[list_JMH_FB["Id_Document_Split"].str.contains("JMH", na=False)]
        list_JMH_FB = list_JMH_FB[list_JMH_FB["Id_Document_Split"].str.contains(regex_kode, regex=True, na=False)]
        list_JMH_FB.rename(columns={"Id_Document_Split": "No Dokumen"}, inplace=True)
        list_JMH_FB["sumber"] = "FB"

        JMH_gabungan = pd.concat([list_JMH_FL, list_JMH_FB], ignore_index=True)
        JMH_gabungan = (
            JMH_gabungan
            .drop_duplicates(subset=["No Dokumen", "vesvoy"])
            .sort_values(by=["Port Id", "No Dokumen"])
            .reset_index(drop=True)
        )
        JMH_gabungan = JMH_gabungan[JMH_gabungan["Port Id"].isin(full_list)]
        
        return dfs_FL, dfs_FB, JMH_gabungan, FULL_NO_JMH, full_list

    except Exception as e:
        st.error(f"Terjadi error saat menjalankan algoritma: {e}")
        st.exception(e)
        return None, None


# ====================================================================
# II. SIDEBAR (Parameter Only)
# ====================================================================
with st.sidebar:
    st.title("⚙️ Parameter")
    
    st.subheader("1. Periode Data")
    bulan_options = [b.upper() for b in list_bulan.keys()]
    
    input_bulan = st.selectbox(
        "Pilih Bulan",
        options=bulan_options,
        index=bulan_options.index("SEPTEMBER") if "SEPTEMBER" in bulan_options else 0
    )
    
    input_tahun = st.text_input(
        "Tahun",
        value="2025",
        max_chars=4,
        help="Format: YYYY (contoh: 2025)"
    )

    st.subheader("2. Pilihan Cabang")
    all_cabang_keys = list(all_cabang_dict.keys())
    selected_cabang = st.multiselect(
        "Pilih Cabang (Port Id)",
        options=all_cabang_keys,
        help="Pilih satu atau lebih cabang yang akan diproses."
    )

# ====================================================================
# III. MAIN AREA (Judul, Petunjuk, Upload, Proses)
# ====================================================================

# 1. Judul Halaman
st.title("Pemrosesan KBM Accrual")

# 2. Petunjuk Penggunaan (Di Tengah / Expander)
with st.expander("ℹ️ PETUNJUK PENGGUNAAN & STRUKTUR FILE (PENTING)", expanded=True):
    st.markdown("""
    **Langkah-langkah:**
    1. Upload file Excel pada bagian di bawah ini.
    2. Atur **Periode** dan **Cabang** pada menu di sebelah kiri (Sidebar).
    3. Klik tombol **Jalankan Pemrosesan Data**.

    ---
    **⚠️ PERHATIAN: Pastikan Susunan Sheet Excel `DATA_KBM.xls` sebagai berikut:**
    1.  **Sheet 1:** KBM Bulan - 2 (Format Lama)
    2.  **Sheet 2:** KBM Bulan - 1 (Format Lama)
    3.  **Sheet 3:** KBM Periode Saat Ini (Format Lama)
    4.  **Sheet 4:** KBM Format Baru (FB)
    5.  **Sheet 5:** File ABM (Opsional/Tidak Wajib)
    """)

# 3. Area Upload File
st.subheader("📂 Upload Data KBM")
uploaded_file = st.file_uploader(
    "Unggah File DATA_KBM.xls",
    type=['xls', 'xlsx']
)

# 4. Tombol Proses
st.divider()
process_button = st.button("🚀 Jalankan Pemrosesan Data", use_container_width=True)

# ====================================================================
# IV. LOGIKA EKSEKUSI & TAMPILAN OUTPUT
# ====================================================================

if process_button:
    # Memanggil fungsi run_processing dengan parameter
    dfs_FL_result, dfs_FB_result, JMH_gabungan_result, FULL_NO_JMH_result, full_list_result = run_processing(
        uploaded_file, input_bulan, input_tahun, selected_cabang, list_bulan, all_cabang_dict, list_port_full, list_cy_full
    )

    if dfs_FL_result is not None:
        st.success("✅ Pemrosesan Data Selesai! File siap diunduh.")

        # Menulis ke memori (buffer) untuk diunduh
        output = io.BytesIO()
        output_file_name = f"OUTPUT_KBM_{input_bulan}_{input_tahun}.xlsx"
        
        # 1. TULIS EXCEL TANPA STYLING ke buffer
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # --- Tulis tiap port ---
            for port in full_list_result:
                df_fl = dfs_FL_result.get(port, pd.DataFrame())
                df_fb = dfs_FB_result.get(port, pd.DataFrame())

                df_fl.to_excel(writer, sheet_name=port, index=False, startrow=0)

                start_fb = len(df_fl) + 4
                df_fb.to_excel(writer, sheet_name=port, index=False, startrow=start_fb)

            # --- Tambah sheet List JMH ---
            JMH_gabungan_result.to_excel(writer, sheet_name="List JMH", index=False)

            # --- Tambah sheet JURNAL NO JMH ---
            FULL_NO_JMH_result.to_excel(writer, sheet_name="JURNAL NO JMH", index=False, header=False)

        # Simpan buffer untuk dimuat ulang (agar bisa diberi styling)
        output.seek(0)
        
        # 2. TAMBAHKAN GARIS PEMBATAS (Styling OpenPyXL)
        try:
            wb = load_workbook(output)
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            for port in full_list_result:
                if port in wb.sheetnames:
                    ws = wb[port]
                    df_fl = dfs_FL_result.get(port, pd.DataFrame())
                    separator_row = len(df_fl) + 3
                    max_col = ws.max_column

                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=separator_row, column=col)
                        cell.fill = black_fill
            
            # Simpan workbook yang sudah distyling ke buffer baru
            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            # Tampilkan tombol download
            st.download_button(
                label="📥 Unduh Hasil Pemrosesan (Excel)",
                data=final_output,
                file_name=output_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            # PREVIEW DIHAPUS SESUAI PERMINTAAN

        except Exception as e:
            st.warning(f"Error saat menambahkan styling. Mengunduh file tanpa styling. Error: {e}")
            output.seek(0)
            st.download_button(
                label="📥 Unduh Hasil Pemrosesan (Tanpa Styling)",
                data=output,
                file_name=output_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )